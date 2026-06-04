import os
import openpyxl

excel_path = r"c:\Users\86135\Desktop\面试题\广东嵌入式Linux岗位分析报告.xlsx"
out_path = r"C:\Users\86135\.gemini\antigravity\brain\d8c67ea2-5201-45be-b279-f1f8a1e5ec1b\scratch\excel_content.txt"

if not os.path.exists(excel_path):
    print("File not found")
else:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    with open(out_path, "w", encoding="utf-8") as f:
        for sheet_name in wb.sheetnames:
            f.write(f"=== Sheet: {sheet_name} ===\n")
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                # Filter out rows that are entirely None
                if any(x is not None for x in row):
                    # Replace None with empty string for cleaner printing
                    row_str = ", ".join(str(x) if x is not None else "" for x in row)
                    f.write(row_str + "\n")
            f.write("\n")
    print("Done")
